"""
文档转换服务
支持：doc/xls/ppt 等旧格式 → docx/xlsx/pptx/pdf
     Excel/CSV → JSON（用于模板导入）
依赖：LibreOffice（headless）、openpyxl、xlrd（可选）
"""
import os
import sys
import json
import shutil
import subprocess
import tempfile
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# 支持的转换关系
# ─────────────────────────────────────────────
CONVERSION_MAP: Dict[Tuple[str, str], str] = {
    # Word
    ("doc",  "docx"): "docx",
    ("doc",  "pdf"):  "pdf",
    ("docx", "pdf"):  "pdf",
    ("docx", "doc"):  "doc",
    ("odt",  "docx"): "docx",
    ("odt",  "pdf"):  "pdf",
    # Excel
    ("xls",  "xlsx"): "xlsx",
    ("xls",  "pdf"):  "pdf",
    ("xlsx", "xls"):  "xls",
    ("xlsx", "pdf"):  "pdf",
    ("ods",  "xlsx"): "xlsx",
    ("ods",  "pdf"):  "pdf",
    # PowerPoint
    ("ppt",  "pptx"): "pptx",
    ("ppt",  "pdf"):  "pdf",
    ("pptx", "ppt"):  "ppt",
    ("pptx", "pdf"):  "pdf",
    ("odp",  "pptx"): "pptx",
    ("odp",  "pdf"):  "pdf",
    # Text
    ("txt",  "pdf"):  "pdf",
    ("md",   "pdf"):  "pdf",
}

SUPPORTED_INPUT_EXTS = {
    "doc", "docx", "odt",
    "xls", "xlsx", "ods", "csv",
    "ppt", "pptx", "odp",
    "txt", "md",
}

SUPPORTED_OUTPUT_EXTS = {"docx", "xlsx", "pptx", "pdf", "xls", "ppt", "doc", "json"}


# ─────────────────────────────────────────────
# LibreOffice 检测
# ─────────────────────────────────────────────
def find_soffice() -> Optional[str]:
    """查找 LibreOffice soffice 可执行文件"""
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/opt/libreoffice/program/soffice",
    ]
    env_path = os.environ.get("LIBREOFFICE_PATH")
    if env_path:
        candidates.insert(0, env_path)

    for p in candidates:
        if os.path.isfile(p):
            return p

    # 尝试 where / which
    try:
        cmd = "where soffice" if sys.platform == "win32" else "which soffice"
        result = subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL)
        path = result.decode().strip().splitlines()[0]
        if path and os.path.isfile(path):
            return path
    except Exception:
        pass

    return None


SOFFICE_PATH: Optional[str] = find_soffice()


# ─────────────────────────────────────────────
# LibreOffice 转换
# ─────────────────────────────────────────────
def convert_with_libreoffice(
    input_path: str,
    output_format: str,
    output_dir: str,
    timeout: int = 300,
) -> Tuple[bool, str]:
    """
    调用 LibreOffice headless 进行文档转换。
    返回 (success, output_file_path_or_error_message)
    """
    if not SOFFICE_PATH:
        return False, "未检测到 LibreOffice 安装，请先安装或设置 LIBREOFFICE_PATH 环境变量"

    work_dir = tempfile.mkdtemp(prefix="kflower_conv_")
    try:
        # 复制输入文件到工作目录，避免权限问题
        src_name = os.path.basename(input_path)
        temp_input = os.path.join(work_dir, src_name)
        shutil.copy2(input_path, temp_input)

        os.makedirs(output_dir, exist_ok=True)

        cmd = [
            SOFFICE_PATH,
            "--headless",
            "--norestore",
            "--nodefault",
            "--nologo",
            "--convert-to", output_format,
            "--outdir", work_dir,
            temp_input,
        ]

        kwargs: Dict[str, Any] = {}
        if sys.platform == "win32":
            from subprocess import CREATE_NO_WINDOW
            kwargs["creationflags"] = CREATE_NO_WINDOW

        proc = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout,
            **kwargs,
        )

        # 找到生成文件
        base_name = Path(src_name).stem
        generated = [
            f for f in os.listdir(work_dir)
            if f.startswith(base_name) and f != src_name
        ]

        if not generated:
            stderr_msg = proc.stderr.decode("utf-8", errors="ignore")
            return False, f"LibreOffice 未生成输出文件。stderr: {stderr_msg[:200]}"

        out_file = generated[0]
        src_full = os.path.join(work_dir, out_file)
        dst_full = os.path.join(output_dir, out_file)

        # 目标已存在则覆盖
        if os.path.exists(dst_full):
            os.remove(dst_full)
        shutil.move(src_full, dst_full)

        return True, dst_full

    except subprocess.TimeoutExpired:
        return False, f"转换超时（>{timeout}s），请检查文件是否损坏"
    except Exception as e:
        return False, f"转换异常: {str(e)}"
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


# ─────────────────────────────────────────────
# Excel / CSV → JSON
# ─────────────────────────────────────────────
def excel_to_json(
    input_path: str,
    sheet_name: Optional[str] = None,
    header_row: int = 0,
    max_rows: int = 5000,
) -> Dict[str, Any]:
    """
    将 xlsx/xls/csv 转换为 JSON 结构。
    返回 {success, data: [{...}], columns: [...], sheet, row_count}
    """
    suffix = Path(input_path).suffix.lower()
    try:
        if suffix == ".csv":
            import csv
            rows = []
            with open(input_path, "r", encoding="utf-8-sig", errors="replace") as f:
                reader = csv.DictReader(f)
                columns = reader.fieldnames or []
                for i, row in enumerate(reader):
                    if i >= max_rows:
                        break
                    rows.append(dict(row))
            return {
                "success": True,
                "data": rows,
                "columns": list(columns),
                "sheet": "Sheet1",
                "row_count": len(rows),
            }

        try:
            import openpyxl
            wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            all_rows = list(ws.iter_rows(values_only=True))
        except Exception:
            # 降级：尝试 xlrd（支持 .xls）
            import xlrd
            wb = xlrd.open_workbook(input_path)
            ws_obj = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
            all_rows = [ws_obj.row_values(r) for r in range(ws_obj.nrows)]
            ws = type("_ws", (), {"title": ws_obj.name})()

        if not all_rows:
            return {"success": True, "data": [], "columns": [], "sheet": "", "row_count": 0}

        # 确定表头行
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(all_rows[header_row])]
        data_rows = all_rows[header_row + 1: header_row + 1 + max_rows]

        result_data = []
        for row in data_rows:
            obj = {}
            for j, h in enumerate(headers):
                val = row[j] if j < len(row) else None
                # 序列化为 Python 基本类型
                if val is None:
                    obj[h] = None
                elif hasattr(val, "isoformat"):
                    obj[h] = val.isoformat()
                else:
                    obj[h] = val
            result_data.append(obj)

        sheet_name_out = getattr(ws, "title", "") or ""
        return {
            "success": True,
            "data": result_data,
            "columns": headers,
            "sheet": sheet_name_out,
            "row_count": len(result_data),
        }

    except Exception as e:
        return {"success": False, "error": str(e)}


# ─────────────────────────────────────────────
# xls → xlsx（纯 Python，无需 LibreOffice）
# ─────────────────────────────────────────────
def xls_to_xlsx(input_path: str, output_path: Optional[str] = None) -> Tuple[bool, str]:
    """
    将 .xls 转换为 .xlsx（依赖 xlrd + openpyxl）。
    """
    try:
        import xlrd
        import openpyxl

        wb_in = xlrd.open_workbook(input_path, formatting_info=False)

        if output_path is None:
            output_path = str(Path(input_path).with_suffix(".xlsx"))

        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)  # 移除默认 sheet

        for sheet_idx in range(wb_in.nsheets):
            ws_in = wb_in.sheet_by_index(sheet_idx)
            ws_out = wb_out.create_sheet(title=ws_in.name)
            for row_idx in range(ws_in.nrows):
                row_data = ws_in.row_values(row_idx)
                ws_out.append([v if v != "" else None for v in row_data])

        wb_out.save(output_path)
        return True, output_path

    except ImportError as e:
        return False, f"缺少依赖库: {e}。请安装 xlrd 和 openpyxl"
    except Exception as e:
        return False, str(e)


# ─────────────────────────────────────────────
# 统一转换入口
# ─────────────────────────────────────────────
def convert_document(
    input_path: str,
    target_format: str,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    统一文档转换入口。
    Args:
        input_path: 源文件绝对路径
        target_format: 目标格式，如 "xlsx" / "pdf" / "docx" / "json"
        output_dir: 输出目录，默认与源文件同目录
    Returns:
        {success, output_path, message, ...}
    """
    input_path = os.path.abspath(input_path)
    if not os.path.isfile(input_path):
        return {"success": False, "error": f"文件不存在: {input_path}"}

    src_ext = Path(input_path).suffix.lstrip(".").lower()
    target_format = target_format.lower().lstrip(".")

    if output_dir is None:
        output_dir = str(Path(input_path).parent)

    # ── JSON 提取（特殊路径）──
    if target_format == "json":
        if src_ext not in ("xlsx", "xls", "ods", "csv"):
            return {"success": False, "error": f"不支持将 .{src_ext} 转换为 JSON"}
        # 若是 xls 先转 xlsx 再提取
        actual_path = input_path
        if src_ext == "xls":
            tmp_xlsx = os.path.join(tempfile.mkdtemp(), Path(input_path).stem + ".xlsx")
            ok, msg = xls_to_xlsx(input_path, tmp_xlsx)
            if not ok:
                # 降级：直接用 xlrd 读
                pass
            else:
                actual_path = tmp_xlsx
        result = excel_to_json(actual_path)
        if result.get("success"):
            # 保存 JSON 文件
            json_path = os.path.join(output_dir, Path(input_path).stem + ".json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(result["data"], f, ensure_ascii=False, indent=2)
            result["output_path"] = json_path
        return result

    # ── xls → xlsx（纯 Python 优先，速度快）──
    if src_ext == "xls" and target_format == "xlsx":
        out_path = os.path.join(output_dir, Path(input_path).stem + ".xlsx")
        ok, msg = xls_to_xlsx(input_path, out_path)
        if ok:
            return {"success": True, "output_path": out_path, "message": "xls → xlsx 转换成功"}
        # 降级到 LibreOffice
        logger.warning(f"xls_to_xlsx 失败 ({msg})，尝试 LibreOffice")

    # ── 检查转换关系 ──
    if (src_ext, target_format) not in CONVERSION_MAP:
        return {
            "success": False,
            "error": f"不支持 .{src_ext} → .{target_format} 的转换",
            "supported_conversions": [f"{s} → {t}" for s, t in CONVERSION_MAP.keys()],
        }

    lo_format = CONVERSION_MAP[(src_ext, target_format)]
    ok, result = convert_with_libreoffice(input_path, lo_format, output_dir)
    if ok:
        return {"success": True, "output_path": result, "message": f"{src_ext} → {target_format} 转换成功"}
    else:
        return {"success": False, "error": result}


# ─────────────────────────────────────────────
# 自动识别 + 转换（上传时钩子使用）
# ─────────────────────────────────────────────
def auto_convert_for_upload(
    input_path: str,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    根据文件扩展名自动转换为「可识别」格式：
      .xls  → .xlsx
      .doc  → .docx
      .ppt  → .pptx
      其他已支持格式直接返回
    用于知识库/模板文件上传时自动处理不兼容格式。
    """
    src_ext = Path(input_path).suffix.lstrip(".").lower()

    auto_targets = {
        "xls": "xlsx",
        "doc": "docx",
        "ppt": "pptx",
        "odt": "docx",
        "ods": "xlsx",
        "odp": "pptx",
    }

    if src_ext not in auto_targets:
        return {
            "success": True,
            "converted": False,
            "output_path": input_path,
            "message": f"格式 .{src_ext} 无需转换",
        }

    target_fmt = auto_targets[src_ext]
    result = convert_document(input_path, target_fmt, output_dir)
    if result.get("success"):
        result["converted"] = True
    return result


# ─────────────────────────────────────────────
# 检测 LibreOffice 状态（供 API 返回）
# ─────────────────────────────────────────────
def get_converter_status() -> Dict[str, Any]:
    """获取文档转换服务状态"""
    has_lo = SOFFICE_PATH is not None

    # 检测 Python 依赖
    deps: Dict[str, bool] = {}
    for lib in ("openpyxl", "xlrd", "PyPDF2", "docx"):
        try:
            __import__(lib)
            deps[lib] = True
        except ImportError:
            deps[lib] = False

    supported = [f"{s} → {t}" for s, t in CONVERSION_MAP.keys()]
    supported.append("xlsx/xls/csv → json")

    return {
        "libreoffice_available": has_lo,
        "libreoffice_path": SOFFICE_PATH,
        "python_dependencies": deps,
        "supported_conversions": supported,
        "ready": has_lo or deps.get("openpyxl", False),
    }
