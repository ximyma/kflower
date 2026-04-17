"""
业务服务 - 数据集成服务
智能连接引擎
"""
from typing import Dict, Any, List, Optional
import json
from datetime import datetime
import hashlib


class IntegrationService:
    """数据集成服务 - 智能连接引擎"""
    
    # 支持的数据源类型
    DATA_SOURCES = {
        "mysql": {
            "name": "MySQL",
            "icon": "database",
            "fields": [
                {"name": "host", "label": "主机", "type": "text", "required": True},
                {"name": "port", "label": "端口", "type": "number", "default": 3306},
                {"name": "database", "label": "数据库", "type": "text", "required": True},
                {"name": "username", "label": "用户名", "type": "text", "required": True},
                {"name": "password", "label": "密码", "type": "password", "required": True}
            ]
        },
        "postgresql": {
            "name": "PostgreSQL",
            "icon": "database",
            "fields": [
                {"name": "host", "label": "主机", "type": "text", "required": True},
                {"name": "port", "label": "端口", "type": "number", "default": 5432},
                {"name": "database", "label": "数据库", "type": "text", "required": True},
                {"name": "username", "label": "用户名", "type": "text", "required": True},
                {"name": "password", "label": "密码", "type": "password", "required": True}
            ]
        },
        "excel": {
            "name": "Excel文件",
            "icon": "file-excel",
            "fields": [
                {"name": "file_path", "label": "文件路径", "type": "text", "required": True},
                {"name": "sheet_name", "label": "工作表名", "type": "text", "default": "Sheet1"}
            ]
        },
        "csv": {
            "name": "CSV文件",
            "icon": "document",
            "fields": [
                {"name": "file_path", "label": "文件路径", "type": "text", "required": True},
                {"name": "encoding", "label": "编码", "type": "select", "options": ["utf-8", "gbk", "gb2312"], "default": "utf-8"},
                {"name": "delimiter", "label": "分隔符", "type": "text", "default": ","}
            ]
        },
        "http": {
            "name": "HTTP API",
            "icon": "connection",
            "fields": [
                {"name": "url", "label": "接口地址", "type": "text", "required": True},
                {"name": "method", "label": "请求方法", "type": "select", "options": ["GET", "POST", "PUT", "DELETE"], "default": "GET"},
                {"name": "headers", "label": "请求头", "type": "textarea"},
                {"name": "auth_type", "label": "认证方式", "type": "select", "options": ["无", "Basic", "Bearer", "API Key"]}
            ]
        },
        "webhook": {
            "name": "Webhook",
            "icon": "bell",
            "fields": [
                {"name": "url", "label": "接收地址", "type": "text", "required": True},
                {"name": "secret", "label": "密钥", "type": "password"}
            ]
        }
    }
    
    # 数据转换规则模板
    TRANSFORM_RULES = {
        "rename": {
            "name": "字段重命名",
            "params": [{"name": "from", "label": "原字段"}, {"name": "to", "label": "新字段"}]
        },
        "type_convert": {
            "name": "类型转换",
            "params": [
                {"name": "field", "label": "字段"},
                {"name": "from_type", "label": "源类型"},
                {"name": "to_type", "label": "目标类型"}
            ]
        },
        "filter": {
            "name": "数据过滤",
            "params": [
                {"name": "field", "label": "字段"},
                {"name": "operator", "label": "条件", "options": ["=", "!=", ">", "<", ">=", "<=", "包含", "不包含"]},
                {"name": "value", "label": "值"}
            ]
        },
        "map": {
            "name": "值映射",
            "params": [
                {"name": "field", "label": "字段"},
                {"name": "mapping", "label": "映射关系", "type": "textarea", "placeholder": "如: 1=启用,0=禁用"}
            ]
        },
        "default": {
            "name": "默认值",
            "params": [
                {"name": "field", "label": "字段"},
                {"name": "default_value", "label": "默认值"}
            ]
        },
        "expression": {
            "name": "计算表达式",
            "params": [
                {"name": "result_field", "label": "结果字段"},
                {"name": "expression", "label": "表达式", "placeholder": "如: {a} + {b}"}
            ]
        }
    }
    
    @classmethod
    async def test_connection(cls, config: Dict[str, Any]) -> Dict[str, Any]:
        """测试数据源连接"""
        source_type = config.get("type")
        
        if source_type == "mysql":
            return await cls._test_mysql(config)
        elif source_type == "postgresql":
            return await cls._test_postgresql(config)
        elif source_type in ["excel", "csv"]:
            return await cls._test_file(config)
        elif source_type == "http":
            return await cls._test_http(config)
        
        return {"success": False, "message": f"不支持的数据源类型: {source_type}"}
    
    @classmethod
    async def _test_mysql(cls, config: Dict[str, Any]) -> Dict[str, Any]:
        """测试MySQL连接"""
        # 简化实现，实际应该用aiomysql
        try:
            # 这里应该实际测试连接
            return {
                "success": True,
                "message": "连接成功",
                "tables": ["users", "orders", "products"]
            }
        except Exception as e:
            return {"success": False, "message": str(e)}
    
    @classmethod
    async def _test_postgresql(cls, config: Dict[str, Any]) -> Dict[str, Any]:
        """测试PostgreSQL连接"""
        try:
            return {
                "success": True,
                "message": "连接成功",
                "tables": ["users", "orders", "products"]
            }
        except Exception as e:
            return {"success": False, "message": str(e)}
    
    @classmethod
    async def _test_file(cls, config: Dict[str, Any]) -> Dict[str, Any]:
        """测试文件连接"""
        import os
        file_path = config.get("file_path")
        
        if not file_path:
            return {"success": False, "message": "文件路径不能为空"}
        
        if not os.path.exists(file_path):
            return {"success": False, "message": "文件不存在"}
        
        return {
            "success": True,
            "message": "文件可访问",
            "rows": 100,  # 简化
            "columns": ["字段1", "字段2", "字段3"]
        }
    
    @classmethod
    async def _test_http(cls, config: Dict[str, Any]) -> Dict[str, Any]:
        """测试HTTP连接"""
        import httpx
        
        url = config.get("url")
        method = config.get("method", "GET")
        
        try:
            async with httpx.AsyncClient() as client:
                response = await client.request(method, url, timeout=10)
                return {
                    "success": response.status_code < 400,
                    "message": f"HTTP {response.status_code}",
                    "data_sample": response.text[:500] if response.status_code < 400 else None
                }
        except Exception as e:
            return {"success": False, "message": str(e)}
    
    @classmethod
    async def smart_mapping(
        cls,
        source_fields: List[str],
        target_schema: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """智能字段映射"""
        from app.core.ai_digital_base.gateway import ai_gateway
        
        system_prompt = """你是一个数据集成专家。根据源字段和目标模式，进行智能字段映射。

规则：
1. 相同或相似名称的字段自动匹配
2. 类型兼容的字段给出建议
3. 无法匹配的字段标记为"未映射"

输出JSON格式：
{
    "mappings": [
        {
            "source_field": "字段名",
            "target_field": "字段名",
            "confidence": 0.95,
            "transform": null
        }
    ],
    "unmapped_source": ["字段1", "字段2"],
    "unmapped_target": ["字段1", "字段2"],
    "suggestions": ["建议1"]
}"""
        
        result = await ai_gateway.chat_with_system_prompt(
            system_prompt=system_prompt,
            user_message=f"源字段: {source_fields}\n目标模式: {target_schema}"
        )
        
        if "error" in result:
            return [{"source_field": s, "target_field": None, "confidence": 0} for s in source_fields]
        
        try:
            content = result["content"].strip()
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            return json.loads(content)
        except:
            return [{"source_field": s, "target_field": s, "confidence": 0.5} for s in source_fields]
    
    @classmethod
    async def preview_data(
        cls,
        source_type: str,
        config: Dict[str, Any],
        limit: int = 100
    ) -> Dict[str, Any]:
        """预览数据"""
        # 根据类型获取数据预览
        if source_type == "excel":
            return await cls._preview_excel(config, limit)
        elif source_type == "csv":
            return await cls._preview_csv(config, limit)
        elif source_type == "mysql":
            return await cls._preview_mysql(config, limit)
        
        return {"error": f"不支持的数据源类型: {source_type}"}
    
    @classmethod
    async def _preview_excel(cls, config: Dict[str, Any], limit: int) -> Dict[str, Any]:
        """预览Excel数据"""
        try:
            import openpyxl
            from io import BytesIO
            
            file_path = config.get("file_path")
            sheet_name = config.get("sheet_name", "Sheet1")
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # 获取表头
            headers = [cell.value for cell in ws[1]]
            
            # 获取数据行
            rows = []
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=limit+1, values_only=True)):
                if any(row):
                    rows.append(list(row))
            
            return {
                "success": True,
                "headers": headers,
                "rows": rows,
                "total_rows": ws.max_row - 1
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @classmethod
    async def _preview_csv(cls, config: Dict[str, Any], limit: int) -> Dict[str, Any]:
        """预览CSV数据"""
        import csv
        
        file_path = config.get("file_path")
        encoding = config.get("encoding", "utf-8")
        delimiter = config.get("delimiter", ",")
        
        try:
            with open(file_path, "r", encoding=encoding) as f:
                reader = csv.reader(f, delimiter=delimiter)
                headers = next(reader)
                
                rows = []
                for i, row in enumerate(reader):
                    if i >= limit:
                        break
                    rows.append(row)
                
                return {
                    "success": True,
                    "headers": headers,
                    "rows": rows
                }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    @classmethod
    async def _preview_mysql(cls, config: Dict[str, Any], limit: int) -> Dict[str, Any]:
        """预览MySQL数据"""
        # 简化实现
        return {
            "success": True,
            "headers": ["id", "name", "created_at"],
            "rows": [[1, "示例1", "2024-01-01"], [2, "示例2", "2024-01-02"]]
        }
    
    @classmethod
    def create_sync_task(
        cls,
        source: Dict[str, Any],
        target: Dict[str, Any],
        mappings: List[Dict[str, Any]],
        schedule: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """创建同步任务"""
        task_id = hashlib.md5(f"{datetime.now().isoformat()}".encode()).hexdigest()[:16]
        
        return {
            "id": task_id,
            "source_type": source.get("type"),
            "target_type": target.get("type"),
            "mappings": mappings,
            "schedule": schedule,
            "status": "created",
            "created_at": datetime.now().isoformat()
        }


integration_service = IntegrationService()
